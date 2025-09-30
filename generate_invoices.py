#!/usr/bin/env python3
"""
Generate a single multi-page PDF of invoices using your Excel template (config.xlsx).

Template mapping (from your uploaded file):
- Order #: B2
- Order Date: G5 (label in F5)
- Billing block:
    B3: Billing Name
    B4: Address1 (Address2 appended as a new line inside the same cell if present)
    B5: City, State ZIP
- Line items (start at row 8):
    B: Qty, C: Description (wrapped), F: Unit Price, G: Line Total
- Totals labels: H37 "Subtotal", H38 "Sales Tax", H39 "Total Paid"
  Amounts go into the cell to the right (I-column)

Highlighting System (user-friendly text matching):
  Column A: Text to match (case-insensitive, partial matching)
  Column B: Color name or hex code
  
  Examples: "bundle" matches "Canada Goldenrod - Solidago lepida, bundle of 5"
           "huckleberry" matches "Evergreen Huckleberry - Vaccinium ovatum, bundle of 5"
           "tree" matches any item containing "tree" in description
  
  Available colors: red, green, blue, yellow, orange, purple, pink, brown, gray/grey
                   light blue, light green, light yellow, light pink, light gray/light grey
                   Or use hex codes like #FF0000 or FF0000

Requirements:
  pip install pandas openpyxl pywin32
  Windows: Microsoft Excel (for PDF export via COM automation)
  Linux/Mac: LibreOffice (for PDF export): `soffice --headless` available on PATH
"""

import os
import io
import shutil
import tempfile
import platform
import gc
from typing import Dict, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as XLImage
from tqdm import tqdm

# Excel COM support removed - using manual PDF export instead

# ---------- CONFIG ----------
ORDERS_CSV  = "orders.csv"
CONFIG_XLSX = "config.xlsx"      # your template; must include 'Invoice' and 'Highlighting'
OUTPUT_XLSX = "generated_invoices.xlsx"
# PDF output removed - use Excel's native "Export as PDF" feature instead

DATA_START_ROW = 8  # B8 is first data row (moved up one row)
CURRENCY_FMT = '"$"#,##0.00'

# Columns in your template for the table (fixed from inspection)
COL_QTY  = column_index_from_string("B")
COL_DESC = column_index_from_string("C")
COL_UNIT = column_index_from_string("F")
COL_LINE = column_index_from_string("G")

# Column widths to keep within one portrait page (tweak if needed)
WIDTHS = {
    COL_QTY:  8,   # B
    COL_DESC: 52,  # C
    COL_UNIT: 12,  # F
    COL_LINE: 14,  # G
}

# ---------- Color Names ----------
# Standard color names that users can use in the Highlights sheet
COLOR_NAMES = {
    'red': 'FF0000',
    'green': '00FF00', 
    'blue': '0000FF',
    'yellow': 'FFFF00',
    'orange': 'FFA500',
    'purple': '800080',
    'pink': 'FFC0CB',
    'brown': 'A52A2A',
    'gray': '808080',
    'grey': '808080',  # Alternative spelling
    'light blue': '87CEEB',
    'light green': '90EE90',
    'light yellow': 'FFFFE0',
    'light pink': 'FFB6C1',
    'light gray': 'D3D3D3',
    'light grey': 'D3D3D3'  # Alternative spelling
}

# ---------- Helpers ----------
def load_orders(path: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip().str.lower()
    return df

def read_highlights(wb) -> Dict[str, str]:
    """Read text match -> color name/hex from sheet 'Highlighting' or 'Highlights'."""
    for name in ("Highlighting", "Highlights"):
        if name in wb.sheetnames:
            ws = wb[name]
            out = {}
            for text_match, color in ws.iter_rows(min_row=2, values_only=True):
                if text_match:
                    text_str = str(text_match).strip()
                    
                    # Skip reference/header rows (lines starting with = or • or # or being color names themselves)
                    if (text_str.startswith('===') or 
                        text_str.startswith('•') or 
                        text_str.startswith('#') or
                        text_str.lower() in ['standard colors:', 'light colors:', 'custom hex:'] or
                        text_str.lower() in COLOR_NAMES):
                        continue
                    
                    c = str(color or "").strip().lower()
                    
                    # Convert color name to hex if it's a recognized color name
                    if c in COLOR_NAMES:
                        hex_color = COLOR_NAMES[c]
                    elif c.startswith("#"):
                        # Remove # prefix from hex codes
                        hex_color = c[1:]
                    elif len(c) == 6 and all(ch in '0123456789abcdef' for ch in c):
                        # Already a hex code without #
                        hex_color = c.upper()
                    else:
                        # Unknown color or empty, skip this entry
                        continue
                    
                    # Store the text match in lowercase for case-insensitive matching
                    out[text_str.lower()] = hex_color
            return out
    return {}

def currency(cell):
    cell.number_format = CURRENCY_FMT

def rightmost_used_col(ws: Worksheet, max_row: int, min_col_letter: str = "B") -> int:
    min_idx = column_index_from_string(min_col_letter)
    max_idx = min_idx
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            if cell.value not in (None, "") and cell.column > max_idx:
                max_idx = cell.column
    return max_idx

def fit_one_page(ws: Worksheet, last_row: int, min_col_letter: str = "B"):
    """Set up page to preserve text formatting while fitting content."""
    # Only fit width, let height be natural to preserve text formatting
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # Don't force height fitting to preserve text sizes
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4)
    ws.print_options.horizontalCentered = True

    last_col_idx = rightmost_used_col(ws, last_row, min_col_letter=min_col_letter)
    ws.print_area = f"{min_col_letter}1:{get_column_letter(last_col_idx)}{last_row}"

# Global cache for image data to avoid closed file issues
_IMAGE_DATA_CACHE = {}

def adjust_anchor_position(anchor, offset_cols=4):
    """Move an anchor position to the right by offset_cols columns."""
    if not anchor or len(anchor) < 2:
        return anchor
    
    # Extract column letters and row number (e.g., "B1" -> "B", "1")
    import re
    match = re.match(r'([A-Z]+)(\d+)', str(anchor).upper())
    if not match:
        return anchor
    
    col_str, row_str = match.groups()
    
    # Convert column letters to number, add offset, convert back
    from openpyxl.utils import column_index_from_string, get_column_letter
    try:
        col_num = column_index_from_string(col_str) + offset_cols
        new_col_str = get_column_letter(col_num)
        return f"{new_col_str}{row_str}"
    except:
        return anchor

def copy_images_from_template(template_ws: Worksheet, target_ws: Worksheet):
    """Re-add images from the template to the target sheet safely (avoid closed-file errors)."""
    global _IMAGE_DATA_CACHE
    imgs = getattr(template_ws, "_images", [])
    
    for i, img in enumerate(imgs):
        try:
            # Cache key based on image properties
            cache_key = f"img_{i}_{id(template_ws)}"
            
            if cache_key not in _IMAGE_DATA_CACHE:
                # First try to get the image data from the embedded binary data
                data = None
                if hasattr(img, 'ref') and hasattr(img.ref, '_data'):
                    data = img.ref._data
                elif hasattr(img, '_data'):
                    if callable(img._data):
                        data = img._data()
                    else:
                        data = img._data
                
                if data:
                    # Store the original anchor object directly
                    original_anchor = getattr(img, 'anchor', None)
                    
                    _IMAGE_DATA_CACHE[cache_key] = {
                        'data': bytes(data),  # make a copy to avoid reference issues
                        'width': 400,  # Increased by 2x
                        'height': 200,  # Increased by 2x
                        'anchor': original_anchor  # Use original anchor object completely
                    }
                elif getattr(img, "path", None) and os.path.exists(img.path):
                    with open(img.path, 'rb') as f:
                        data = f.read()
                    # Store the original anchor object directly
                    original_anchor = getattr(img, 'anchor', None)
                    
                    _IMAGE_DATA_CACHE[cache_key] = {
                        'data': data,
                        'width': 400,  # Increased by 2x
                        'height': 200,  # Increased by 2x
                        'anchor': original_anchor  # Use original anchor object completely
                    }
                elif os.path.exists("logo.png"):
                    with open("logo.png", 'rb') as f:
                        data = f.read()
                    _IMAGE_DATA_CACHE[cache_key] = {
                        'data': data,
                        'width': 400,  # Increased by 2x
                        'height': 200,  # Increased by 2x
                        'anchor': adjust_anchor_position(anchor, offset_cols=4)  # Move right to avoid data
                    }
                else:
                    continue
            
            # Create fresh image from cached data
            cached = _IMAGE_DATA_CACHE[cache_key]
            buf = io.BytesIO(cached['data'])
            buf.seek(0)
            new_img = XLImage(buf)
            new_img._buffer = buf  # keep alive so writer can read it later
            new_img.width = cached['width']
            new_img.height = cached['height']
            
            # Use the original anchor object if available, otherwise default
            if cached['anchor']:
                new_img.anchor = cached['anchor']
            else:
                new_img.anchor = "F1"  # Default fallback
            
            target_ws.add_image(new_img)
            
        except Exception as e:
            print(f"[warn] Could not copy image {i+1}: {e}")
    
    # If no images were found or processed successfully, add logo.png as fallback
    if not imgs and os.path.exists("logo.png"):
        try:
            new_img = XLImage("logo.png")
            new_img.anchor = "F1"  # Default fallback if no template image
            new_img.width = 400   # Increased by 2x
            new_img.height = 200  # Increased by 2x
            target_ws.add_image(new_img)
        except Exception as e:
            print(f"[warn] Could not add logo.png directly: {e}")

def place_amount(ws: Worksheet, label_text: str, amount: float):
    """Find label cell whose text contains label_text (case-insensitive), write amount to the right."""
    needle = label_text.lower()
    for row in ws.iter_rows(min_row=30, max_row=55, min_col=8, max_col=12):  # focus near your totals area
        for cell in row:
            v = str(cell.value or "").strip().lower()
            if needle in v:
                tgt = ws.cell(row=cell.row, column=cell.column + 1)
                tgt.value = float(amount)
                currency(tgt)
                return True
    return False


# PDF conversion functions removed - use Excel's native "Export as PDF" feature instead
# This is much faster and more reliable for large files


# ---------- Main ----------
def main():
    print("📊 Loading order data...")
    df = load_orders(ORDERS_CSV)
    required = [
        "order id", "created at",
        "billing name", "billing address1", "billing address2", "billing city", "billing province", "billing zip",
        "lineitem quantity", "lineitem name", "lineitem price", "lineitem sku",
        "subtotal", "shipping", "taxes", "total",
    ]
    miss = [c for c in required if c not in df.columns]
    if miss:
        raise ValueError(f"CSV missing columns (after normalization): {miss}")

    print("📋 Loading Excel template...")
    wb = load_workbook(CONFIG_XLSX)
    if "Invoice" not in wb.sheetnames:
        raise ValueError("config.xlsx must contain a sheet named 'Invoice' (case-sensitive).")
    template = wb["Invoice"]

    highlights = read_highlights(wb)

    # Get unique order IDs and count for progress bar
    order_groups = list(df.groupby("order id", sort=False))
    total_orders = len(order_groups)
    
    print(f"Processing {total_orders} orders...")
    
    # One invoice per order (duplicate inside same workbook so styles persist)
    for i, (order_id, g) in enumerate(tqdm(order_groups, desc="Generating invoices", unit="invoice", ascii=True)):
        ws = wb.copy_worksheet(template)
        ws.title = f"Invoice_{order_id}"

        # Re-add images (logo)
        copy_images_from_template(template, ws)

        first = g.iloc[0]

        # ----- Header block -----
        # Order number -> B2 (corrected location)
        ws["B2"] = f"Order {order_id}"

        # Order date -> G5 (label in F5) - extract just the date part
        created_at = str(first.get("created at", ""))
        if created_at:
            # Split by space and take first part (date only)
            date_part = created_at.split()[0] if ' ' in created_at else created_at
            ws["G5"] = date_part
        else:
            ws["G5"] = ""

        # Billing block
        ws["B3"] = str(first.get("billing name", "")).strip()

        # Combine address lines 1 and 2 on same line, skip if addr2 is empty or NaN
        addr1 = str(first.get("billing address1", "")).strip()
        addr2_raw = first.get("billing address2", "")
        addr2 = str(addr2_raw).strip() if pd.notna(addr2_raw) else ""
        if addr2 and addr2.lower() != 'nan':
            ws["B4"] = f"{addr1}, {addr2}"
        else:
            ws["B4"] = addr1
        ws["B4"].alignment = Alignment(wrap_text=True)

        city = str(first.get("billing city", "")).strip()
        prov = str(first.get("billing province", "")).strip()
        # Fix zip code decimal point issue
        zipc_raw = first.get("billing zip", "")
        if pd.notna(zipc_raw):
            try:
                # Convert to string and remove decimal if it's a float
                zipc_str = str(zipc_raw).strip()
                if zipc_str.replace('.', '').replace('-', '').isdigit():
                    zipc = str(int(float(zipc_str)))
                else:
                    zipc = zipc_str
            except (ValueError, TypeError):
                zipc = str(zipc_raw).strip()
        else:
            zipc = ""
        zipc = zipc.strip()
        ws["B5"] = f"{city}, {prov} {zipc}".strip(" ,")

        # ----- Column widths to fit a single page -----
        for cidx, width in WIDTHS.items():
            ws.column_dimensions[get_column_letter(cidx)].width = width

        # ----- Line items (start B9/C9/F9/G9) -----
        r = DATA_START_ROW
        for _, it in g.iterrows():
            qty   = it.get("lineitem quantity", "")
            desc  = str(it.get("lineitem name", "") or "")
            unitp = float(pd.to_numeric(it.get("lineitem price"), errors="coerce") or 0.0)
            sku   = str(it.get("lineitem sku") or "")
            line  = float(pd.to_numeric(qty, errors="coerce") or 0) * unitp

            # Create cells and apply consistent alignment
            cqty = ws.cell(row=r, column=COL_QTY,  value=qty)
            cdesc = ws.cell(row=r, column=COL_DESC, value=desc)
            cunit = ws.cell(row=r, column=COL_UNIT, value=unitp)
            cline = ws.cell(row=r, column=COL_LINE, value=line)
            
            # Format currency cells
            currency(cunit)
            currency(cline)

            # Apply alignment: center for qty/unit/line, left-aligned for description
            cqty.alignment = Alignment(horizontal="center", vertical="center")
            cdesc.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
            cunit.alignment = Alignment(horizontal="center", vertical="center")
            cline.alignment = Alignment(horizontal="center", vertical="center")

            # Optional highlight by description text matching (row fill for B/C/F/G)
            desc_lower = desc.lower()
            matched_color = None
            matched_text = None
            for text_match, hexcolor in highlights.items():
                if text_match in desc_lower:
                    matched_color = hexcolor
                    matched_text = text_match
                    break  # Use first match found
            
            if matched_color:
                fill = PatternFill(fill_type="solid", start_color=matched_color, end_color=matched_color)
                cqty.fill = fill
                cdesc.fill = fill
                cunit.fill = fill
                cline.fill = fill

            ws.row_dimensions[r].height = 22
            r += 1

        # ----- Totals (write to the right of labels in H37/H38/H39) -----
        subtotal = float(pd.to_numeric(first.get("subtotal"), errors="coerce") or 0.0)
        shipping = float(pd.to_numeric(first.get("shipping"), errors="coerce") or 0.0)
        taxes    = float(pd.to_numeric(first.get("taxes"),    errors="coerce") or 0.0)
        total    = float(pd.to_numeric(first.get("total"),    errors="coerce") or (subtotal + shipping + taxes))

        place_amount(ws, "subtotal",  subtotal)
        place_amount(ws, "sales tax", taxes) or place_amount(ws, "tax", taxes)
        place_amount(ws, "total paid", total)

        # ----- Fit to one page (width & height) and include all totals on the right -----
        # Find the actual last used row to include any footer content
        last_used_row = ws.max_row
        last_row = max(r, 40, last_used_row)  # include line items, totals, and any footer content
        fit_one_page(ws, last_row, min_col_letter="B")
        
        # Memory cleanup every 5 invoices to prevent hanging
        if (i + 1) % 5 == 0:
            gc.collect()

    # Remove the template and highlighting sheets from the saved output
    if "Invoice" in wb.sheetnames:
        wb.remove(wb["Invoice"])
    for n in ("Highlighting", "Highlights"):
        if n in wb.sheetnames:
            wb.remove(wb[n])

    print("💾 Saving Excel file...")
    print(f"   Workbook contains {len(wb.worksheets)} sheets, this may take a moment...")
    
    # Force garbage collection before saving
    gc.collect()
    
    try:
        wb.save(OUTPUT_XLSX)
        print(f"   ✅ Saved {OUTPUT_XLSX}")
    except PermissionError:
        print(f"❌ Permission denied: Close {OUTPUT_XLSX} if it's open in Excel/LibreOffice")
        print("   Or delete the existing file and try again")
        raise
    except Exception as e:
        print(f"❌ Error saving Excel file: {e}")
        print("   This may be due to memory constraints with large workbooks")
        raise

    print("✅ Excel file created successfully!")
    print(f"📄 To print invoices: Open {OUTPUT_XLSX} in Excel and print (Ctrl+P)")
    print(f"   Select 'Print Entire Workbook' to print all invoices")


if __name__ == "__main__":
    main()

